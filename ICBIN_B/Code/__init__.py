# -*- coding: utf-8 -*-

"""
Created on Thu Jan 18 15:51:47 2024

@author: Andrew R Wilzman
"""

import numpy as np
import argparse
import os
import h5py
import torch
import arw_training_turing as trn
import networks
from tabulate import tabulate
import openpyxl
import re
import matplotlib.pyplot as plt
from scipy.stats import kendalltau
from matplotlib import cm
from matplotlib.colors import Normalize

def save_losses_h5(data_dir, model_name, loop, losses):
    # Ensure directory exists
    os.makedirs(data_dir, exist_ok=True)
    
    loss_data_path = f'{data_dir}ae_{model_name}_{loop}_losses.h5'
    with h5py.File(loss_data_path, 'w') as f:
        f.create_dataset('losses', data=losses)
    print(f'Loss data saved as: {loss_data_path}')
    
def extract_state_lists(state_dict, layer_prefixes):
    state_lists = []

    for layer_prefix in layer_prefixes:
        layer_list = [key for key in state_dict.keys() if layer_prefix in key and 'weight' in key]
        state = []

        for idx, key in enumerate(layer_list):
            weight_size = state_dict[key].size()
            state.append([weight_size[1], weight_size[0]])

        state_lists.append(state)

    return state_lists

def grow_network(network, losses, thresh_scale=0.8, width_scale=0.2):
    losses = torch.tensor(losses).detach().cpu()
    quarter_size = len(losses) // 4
    first_quarter_losses = losses[quarter_size:2 * quarter_size]
    third_quarter_losses = losses[2 * quarter_size:3 * quarter_size]
    fourth_quarter_losses = losses[3 * quarter_size:]

    # Calculate mean losses for each quarter
    first_quarter_max = first_quarter_losses.max()
    third_quarter_mean = third_quarter_losses.mean()
    fourth_quarter_mean = fourth_quarter_losses.mean()

    cond = ((fourth_quarter_mean > third_quarter_mean * thresh_scale) and
            (fourth_quarter_mean < third_quarter_mean / thresh_scale) and
            (fourth_quarter_mean < first_quarter_max * thresh_scale))
    if cond:
        print('Growing the network!')
        for layer_type in range(1, 5):
            feat_list = []
            
            for mo in network.get_layer_list(layer_type):
                feat_list.extend([mo.in_features, mo.out_features])
                
            a = feat_list[0]
            b = feat_list[-1] if len(feat_list) < 4 else feat_list[3]
            
            new_width = int(a * width_scale + b * (1 - width_scale))
            
            network.add_layer(layer_type, 0)
            network.change_width(layer_type, 0, new_width)
        
    return network
                
def gather_samples(data, filter_by_cdvr, fatigue_map, num_points=2048):
    selected_samples = {}
    pattern = re.compile(r'^\d{7}[A-Za-z]?$')  # Your regex pattern

    for file_path in data:
        # Extract the folder name from the file path
        folder_name = os.path.basename(os.path.dirname(file_path))  # Get folder name

        # Apply the filter by checking the folder name
        if filter_by_cdvr:
            if not pattern.match(folder_name):
                continue

        try:
            with h5py.File(file_path, 'r') as hf:
                # Check if required datasets exist
                if 'Surface' not in hf or 'MTno' not in hf or 'Side' not in hf:
                    continue  # Skip this file if any required dataset is missing

                bone = hf['Surface'][:]  # (num_points, 3)
                if bone.shape[0] == 0:
                    continue  # Skip this file if no points in the surface dataset
                
                mtno = int(np.array(hf['MTno']))
                side = str(np.array(hf['Side'])[()].decode())

            # Sample the points
            n = bone.shape[0]
            sampled_indices = np.random.choice(n, min(num_points, n), replace=False)
            sampled_bone = bone[sampled_indices]  # shape: (min(num_points, n), 3)

            key = f'{folder_name}{side}{mtno}'

            # Check if fatigue value exists in fatigue_map
            fatigue = fatigue_map.get(key)
            if fatigue is None:
                continue  # Skip if no fatigue value found

            selected_samples[key] = (sampled_bone, fatigue)

        except Exception as e:
            # Optionally, log the exception for debugging
            # print(f"Error processing {file_path}: {e}")
            continue  # Skip any file that raises an error

    # Create X, y by extracting samples and fatigue values
    X = [sample[0] for sample in selected_samples.values()]
    y = [sample[1] for sample in selected_samples.values()]

    # Convert to numpy arrays for consistency
    X = np.array(X)
    y = np.array(y)

    return X, y


#Z:/_PROJECTS/Deep_Learning_HRpQCT/ICBIN_B/
if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument('--direct', type=str,default='')
    
    parser.add_argument('-a','--autoencode', action='store_true')
    parser.add_argument('--vae', action='store_true')
    parser.add_argument('-d','--diffuse', action='store_true')
    parser.add_argument('-g','--gan', action='store_true')
    
    parser.add_argument('--seed', type=int,default=0)
    parser.add_argument('-e','--epochs', type=int,default=0)
    parser.add_argument('-t','--traintime', type=int,default=0)
    parser.add_argument('-i','--init', type=int,default=1)
    parser.add_argument('--grow', action='store_true')
    parser.add_argument('--chpt',type=int,default=0)
    parser.add_argument('--grow_thresh',type=float,default=0.8)
    parser.add_argument('--grow_width',type=float,default=0.2)
    parser.add_argument('-lr', type=float,default=1e-3)
    parser.add_argument('--decay', type=float,default=1e-6)
    parser.add_argument('--chkdecay',type=float,default=0.95)
    parser.add_argument('--batch', type=int,default=1)
    parser.add_argument('--eval_bs', type=int, default=8, help='eval batch size')
    parser.add_argument('--pint', type=int,default=0)
    parser.add_argument('--noise', type=int,default=3)
    parser.add_argument('--hidden1', type=int,default=512)
    parser.add_argument('--hidden2', type=int,default=128)
    parser.add_argument('--hidden3', type=int,default=128)

    parser.add_argument('--name', type=str,default='')
    parser.add_argument('--loadgen', type=str,default='')
    parser.add_argument('--loaddis', type=str,default='')
    parser.add_argument('--pc_gen', type=int,default=0)
    
    parser.add_argument('--cycles', type=int, default=1)
    parser.add_argument('--numpoints', type=int,default=1024)
    parser.add_argument('-v','--visual', action='store_true')
    parser.add_argument('-n','--network', type=str, choices=['trs', 'fold'],
                        help='Network call sign')
    
    args = parser.parse_args(['--direct','../','-n','fold',
                              #'-v',
                              '--seed','0',
                              '--vae',
                              '--batch','32',
                              '--hidden1','512',
                              '-lr','3e-3','--decay','1e-5',
                              '--numpoints','512',
                              '-e','1',
                              '-t','0',
                              '--pint','1',
                              '--chpt','0',
                              '--cycles','1',
                              '--noise','3',
                              '--name','test',
                              '--pc_gen','0',
                              '--loadgen','Vae_test_512_128_128_0',
                              '--loaddis',''])
                    
    #Initialize vars
    if args.loadgen != '':
        if args.loadgen[-4:] != '.pth': # must be .pth
            args.loadgen += '.pth'
    if args.seed == 0:
        args.seed = torch.randint(10, 8545, (1,)).item() 
    num_points = args.numpoints
    directory = args.direct+'Data/'
    epochs = args.epochs
    learning_rate = args.lr
    wtdecay = args.decay
    batch_size = args.batch
    print_interval = args.pint
    cycles = args.cycles    
    
    loops = 1
    if args.chpt > 0:
        loops = args.chpt + 1
        
    if epochs == 0:
        epochs = -args.traintime
            
    if loops > 1:
        epochs = epochs // loops
        
    if print_interval == 0:
        print_interval = max(1,abs(epochs) // 10)
    
    samps = os.listdir(directory+'Compressed/')
    data = []
    
    # Take inventory and read in sample data
    # Remove metatarsals 1 and 5
    xout = ['L1', 'L5', 'R1', 'R5']
    pattern = re.compile(r'^\d{7}[A-Za-z]?$')
    
    for s in samps:
        if not pattern.match(s):
            continue
        s_folder = f'{directory}Compressed/{s}'
        MTs = os.listdir(s_folder)
        for MT in MTs:
            if not any(x in MT for x in xout):
                h5_file = f'{directory}Compressed/{s}/{MT}'
                data.append(h5_file)
                
    # Network setup
    if torch.cuda.is_available():
        print('CUDA available')
        print(torch.cuda.get_device_name(0))
        n_gpus = torch.cuda.device_count()
        print(f"Number of GPUs available: {n_gpus}")
    else:
        print('CUDA *not* available')
    device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')

    torch.manual_seed(args.seed)
    print(f'seed: {args.seed}')
    
    # Initialize state
    if args.network == 'trs':
        input_dim = 3
        input2_dim = args.hidden3+3
    elif args.network == 'fold':
        input_dim = 12
        input2_dim = args.hidden1
        
    if args.network == 'trs':
        network = networks.arw_TRSNet(args.hidden1,args.hidden3).to(device)
    elif args.network == 'fold':
        network = networks.arw_FoldingNet(args.hidden1,args.hidden3).to(device)
        
    if n_gpus > 1:
        network = torch.nn.DataParallel(network)
        
    if args.loadgen != '':
        network.load_state_dict(torch.load(f'{args.direct}Models/{args.loadgen}',weights_only=False))
        
    if args.visual: #see example input
        import open3d as o3d
        import pointcloud_handler as pch
        import pandas as pd
        import matplotlib.pyplot as plt
        from sklearn.decomposition import IncrementalPCA
        
        def set_point_cloud_color(point_cloud, color):
            point_cloud.colors = o3d.utility.Vector3dVector(np.tile(color, (len(point_cloud.points), 1)))
        
        color1 = np.array([0.229, 0.298, 0.512])
        color2 = np.array([0.728, 0.440, 0.145])
        color3 = np.array([0.598, 0.770, 0.344])

        fatigue_map = {}

        # Open the Excel file
        wb = openpyxl.load_workbook(f"{directory}Fatigue/Cadaver_Loadcases.xlsx")
        ws = wb.active  # Get the active sheet

        # Find the column indices for the required fields
        header = [cell.value for cell in ws[1]]  # Assuming first row is header
        idx = [header.index(c) for c in ['Foot', 'Side', 'Mtno', 'Angle', 'Fatigue Life']]

        # Iterate over the rows
        for row in ws.iter_rows(min_row=2, values_only=True):  # Skip the header row
            foot, side, mtno, angle, fatigue = [row[i] for i in idx]
            key = f'{foot}{side}{mtno}'
            fatigue_map[key] = float(fatigue)
        print(f"Fatigue map size: {len(fatigue_map)}")
        print(f"First 5 entries in fatigue_map: {list(fatigue_map.items())[:5]}")
        X, y = gather_samples(data, True, fatigue_map)
        print(f"Number of samples in X: {len(X)}")
        
        all_encodings = []

        # Iterate through each sample and get encoded vector
        for samp in range(len(X)):
            with torch.no_grad():
                # Pass the sample through the network for encoding
                enc, _ = network.encode(torch.FloatTensor(X[samp]).unsqueeze(0).to(device))
                all_encodings.append(enc.squeeze(0).cpu())  # Shape: (2, 512)
        print(f"Number of encodings: {len(all_encodings)}")
        # Stack all encodings into a single tensor (Shape: (N, 2, 512))
        all_encodings = torch.stack(all_encodings)

        # If you need the mean and standard deviation of the encoded vectors (optional)
        mean_enc = all_encodings.mean(dim=0)  # Shape: (2, 512)
        std_enc = all_encodings.std(dim=0)     # Shape: (2, 512)
                
        # Plotting the mean of encoded vectors
        fig, axs = plt.subplots(2, 8, figsize=(16, 4))  # 2 rows, 8 images per row
        for row in range(2):
            vec = mean_enc[row]  # Shape: (512,)
            for i in range(8):
                img = vec[i*64:(i+1)*64].reshape(8, 8).numpy()  # Reshape to (8, 8)
                axs[row, i].imshow(img, cmap='gray')
                axs[row, i].axis('off')
        
        plt.suptitle("Mean of Encoded Vectors",fontsize=24)
        plt.tight_layout()
        plt.show()
        
        # Plotting the std of encoded vectors
        fig, axs = plt.subplots(2, 8, figsize=(16, 4))  # 2 rows, 8 images per row
        for row in range(2):
            vec = std_enc[row]  # Shape: (512,)
            for i in range(8):
                img = vec[i*64:(i+1)*64].reshape(8, 8).numpy()  # Reshape to (8, 8)
                axs[row, i].imshow(img, cmap='gray')
                axs[row, i].axis('off')
        
        plt.suptitle("Standard Deviation of Encoded Vectors",fontsize=24)
        plt.tight_layout()
        plt.show()

        y_array = np.array(y)
        
        N, D1, D2 = all_encodings.shape  # D1=2, D2=512

        # Compute Kendall's tau-b for each latent variable
        tau_vals = np.zeros((D1, D2))
        p_vals = np.zeros((D1, D2))

        for row in range(D1):
            for i in range(D2):
                latent = all_encodings[:, row, i].numpy()
                tau, p = kendalltau(latent, y_array)
                tau_vals[row, i] = tau
                p_vals[row, i] = p

        # Normalize tau values for colormap (map to [0,1])
        norm = Normalize(vmin=np.nanmin(tau_vals), vmax=np.nanmax(tau_vals))
        cmap = cm.Greens

        # Plot with coloring based on Kendall's tau, highlight significance
        fig, axs = plt.subplots(2, 8, figsize=(16, 4))
        for row in range(2):
            for i in range(8):
                idx_start = i * 64
                idx_end = (i + 1) * 64

                tau_patch = tau_vals[row, idx_start:idx_end]
                sig_patch = p_vals[row, idx_start:idx_end] < 0.05  # significance mask
                img = tau_patch.reshape(8, 8)

                color_img = cmap(norm(img))
                # Dim non-significant entries
                color_img[~sig_patch.reshape(8, 8)] = [0.2, 0.2, 0.2, 1.0]  # dark gray

                axs[row, i].imshow(color_img)
                axs[row, i].axis('off')

        plt.suptitle("Kendall's Tau-b: Green = Significant Correlation", fontsize=20)
        plt.tight_layout()
        plt.show()

        sig_mask = p_vals < 0.05

        # Calculate mean and standard deviation of all encodings
        enc_mean = all_encodings.mean(dim=0).cpu().numpy()
        enc_std = all_encodings.std(dim=0).cpu().numpy()

        # Range of interpolation to vary the latent dimensions
        interp_range = np.linspace(-6, 6, 40)  # Adjust this to control the granularity of variations

        # Prepare for visualization
        pcs_all = []  # List to hold all point clouds generated for visualization
        latent_grid = []

        # Define the color map (from red to blue)
        cmap = cm.coolwarm  # You can also try cm.cool or cm.rainbow

        # Iterate over rows and columns of the latent space
        for row in range(2):
            for i in range(512):
                if not sig_mask[row, i]:  # Skip dimensions that aren't significant
                    continue

                # For significant latent dimensions, generate point clouds by varying them within +/- 3 std dev
                for alpha in interp_range:
                    enc_mod = np.copy(enc_mean)  # Copy the mean encoding
                    enc_mod[row, i] += alpha * enc_std[row, i]  # Vary the current latent dimension

                    # Convert to tensor and move to device
                    enc_tensor = torch.FloatTensor(enc_mod).unsqueeze(0).to(device)
                    with torch.no_grad():
                        pc = network.decode(enc_tensor, num_points).squeeze().cpu().numpy()

                    latent_grid.append(pc)

        # Convert the list of point clouds into a single numpy array
        latent_grid = np.vstack(latent_grid)

        # Now assign colors based on the alpha values
        colors = []

        # For each point in the point cloud, map the corresponding alpha value to a color
        for idx, alpha in enumerate(np.tile(interp_range, len(sig_mask))):  # Use np.tile to match number of points
            # Map alpha to a color using the colormap (alpha normalized to [0, 1])
            norm_alpha = (alpha - interp_range.min()) / (interp_range.max() - interp_range.min())  # Normalize alpha
            color = cmap(norm_alpha)[:3]  # Get the RGB part of the colormap
            colors.append(color)

        # Convert to numpy array for Open3D
        colors = np.array(colors)

        # Create Open3D PointCloud object
        pc_o3d = o3d.geometry.PointCloud()
        pc_o3d.points = o3d.utility.Vector3dVector(latent_grid)
        pc_o3d.colors = o3d.utility.Vector3dVector(colors)  # Apply the colors to the point cloud

        # Visualize the point cloud
        o3d.visualization.draw_geometries([pc_o3d], window_name="Latent Space Point Cloud Extrapolation")
                
        # Flatten the encodings for PCA
        flattened_encodings = all_encodings.view(-1, args.hidden1*2).numpy()
        
        # Apply IncrementalPCA to the flattened encodings
        n_components = 3  # Top 3 principal components\
        
        # Initialize IncrementalPCA
        ipca = IncrementalPCA(n_components=n_components)
        ipca.fit(flattened_encodings)
        
        # Now, transform the data with the fitted PCA model
        pca_embedded = ipca.transform(flattened_encodings)
        
        # For visualization with Open3D, let’s display point clouds
        # Generate points for each interpolation range (-3 to +3 for each principal component)
        std_pca = np.std(pca_embedded, axis=0)
        num_interpolated_points = 6  # Number of interpolated points to generate
        interpolated_points = []
        
        # For each principal component, interpolate from -3 to 3
        pca_range = np.linspace(-3, 3, num_interpolated_points)  # Interpolation values for the 3 components
        
        for pca_value_1 in pca_range:
            for pca_value_2 in pca_range:
                for pca_value_3 in pca_range:
                    # Interpolate the principal components
                    interp_pca = np.copy(pca_embedded.mean(axis=0))  # Start with the mean of the embeddings
                    interp_pca[0] += pca_value_1 * std_pca[0]*10  # Modify the first principal component
                    interp_pca[1] += pca_value_2 * std_pca[1]*10  # Modify the second principal component
                    interp_pca[2] += pca_value_3 * std_pca[2]*10 # Modify the third principal component
        
                    # Now we need to reverse the PCA transformation to get back to the original space (shape: (N, 2, 512))
                    # First, revert the PCA to the original space using the inverse transform
                    pca_original_space = ipca.inverse_transform(interp_pca.reshape(1, -1))  # (1, 1024)
        
                    # Reshape back to the format that the decoder expects: (1, 2, 512)
                    interp_pca_resized = pca_original_space.reshape(1, 2, args.hidden1)
                    interp_pca_resized = torch.FloatTensor(interp_pca_resized).to(device)
        
                    # Decode the interpolated embeddings
                    fake_point_cloud = network.decode(interp_pca_resized, num_points)
        
                    # Store the point cloud for visualization
                    point_cloud = o3d.geometry.PointCloud()
                    point_cloud.points = o3d.utility.Vector3dVector(fake_point_cloud.squeeze().cpu().detach().numpy())
                    
                    norm_value = (pca_value_1 + 3) / 6  # Normalize from -3 to +3 to [0, 1] range for color map
                    color_map = plt.cm.coolwarm(norm_value) 
                    # Colorize points based on the interpolation values
                    color_rgb = color_map[:3]
                    set_point_cloud_color(point_cloud, color_rgb)
        
                    interpolated_points.append(point_cloud)
        
        # Visualize the point clouds with Open3D
        o3d.visualization.draw_geometries(interpolated_points)
        
        if args.pc_gen > 0:
            good_set = []
            vnov_set = []
            bad_set = []
            
            # Save point clouds to CSV
            
            for i in range(args.pc_gen):
                noise = (torch.rand_like(enc)-0.5)*2
                with torch.no_grad():
                    fake = network.decode(noise, num_points)
                fake = fake.cpu().detach().numpy()
                
                # Flatten each generated point cloud (num_points, 3)
                points = pd.DataFrame(fake[0], columns=['x', 'y', 'z'])
                #points, _ = pch.inc_PCA(points)
                vnv = pch.create_stl(points,
                                     f'{args.direct}Data/Generated/unprocessed/{args.loadgen[:-4]}/{args.loadgen[:-4]}_{i}',
                                     16,False)
                if vnv > 0:
                    good_set.append(noise)
                elif vnv == 0:
                    vnov_set.append(noise)
                else:
                    bad_set.append(noise)                   
            bad_set = np.array([n.detach().cpu().numpy() for n in bad_set])
            vnov_set = np.array([n.detach().cpu().numpy() for n in vnov_set])
            good_set = np.array([n.detach().cpu().numpy() for n in good_set])
            
    if args.grow:
        perc_thresh = int(args.grow_thresh*100)
        perc_width = int(args.grow_width*100)
        model_name = f'{args.name}_{args.hidden1}_{perc_thresh}_{perc_width}'
    else:
        model_name = f'{args.name}_{args.hidden1}_{args.hidden2}_{args.hidden3}'
    
#%%
    data = []
    
    # Take inventory and read in sample data
    # Remove metatarsals 1 and 5
    xout = ['L1', 'L5', 'R1', 'R5']
    
    for s in samps:
        s_folder = f'{directory}Compressed/{s}'
        MTs = os.listdir(s_folder)
        for MT in MTs:
            if not any(x in MT for x in xout):
                h5_file = f'{directory}Compressed/{s}/{MT}'
                data.append(h5_file)
                
    if args.diffuse or args.autoencode or args.vae:
        loss_function = trn.Chamfer_Loss()
                      
        if args.autoencode:
            if args.diffuse:
                print_interval = max(1,print_interval // 2)
                
            for loop in range(loops):
                network,losses = trn.train_autoencoder(
                    data,network,epochs,learning_rate,wtdecay,batch_size,
                    loss_function,print_interval,device,num_points,cycles)
                
                loss_file = f'{args.direct}Metrics/'
                save_losses_h5(loss_file, model_name, loop, losses)
                
                # save networks                
                torch.save(network.state_dict(),f'{args.direct}Models/ae_{model_name}_{loop}.pth')
                print(f'Model saved as: {args.direct}Models/ae_{model_name}_{loop}.pth')
                
                chf_loss,emd = trn.model_eval_chamfer(data, network.to(device), 
                                                      num_points, device, batch_size=args.eval_bs)
                
                # Create and save metric table 
                table = [["Metric", "Value"],
                         ["Chamfer Loss", f'{chf_loss:.3f}'],
                         ["EMD Loss", f'{emd:.3f}'],
                         ["Epochs", f'{epochs}'],
                         ["Learn Rate", f'{learning_rate}'],
                         ["Decay", f'{args.decay}'],
                         ["Batch Size", f'{args.batch}']]
                
                
                print(tabulate(table, headers="firstrow", tablefmt="fancy_grid", numalign="right"))
                
                cycles += 1
                
                if args.grow:
                    network = grow_network(network,losses,args.grow_thresh,args.grow_width)
                    network = network.to(device)
                    
            learning_rate = learning_rate * args.chkdecay
            filename = f'{args.direct}Metrics/ae_{model_name}.txt'
            with open(filename, 'w', encoding='utf-8') as file:
                file.write(tabulate(table, headers="firstrow", tablefmt="fancy_grid", numalign="right"))
                
        if args.vae:
            cycles = args.cycles
            for loop in range(loops):
                network,losses = trn.train_vae(data,network,epochs, 
                                               learning_rate,wtdecay,batch_size, 
                                               loss_function,print_interval, 
                                               device,num_points)
                torch.save(network.state_dict(),f'{args.direct}Models/Vae_{model_name}_{loop}.pth')
                print(f'Model saved as: {args.direct}Models/Vae_{model_name}_{loop}.pth')
                
                chf_loss,emd = trn.model_eval_chamfer(data, network.to(device),
                                                      num_points, device, batch_size=args.eval_bs)
                
                # Create and save metric table 
                table = [["Metric", "Value"],
                         ["Chamfer Loss", f'{chf_loss:.3f}'],
                         ["EMD Loss", f'{emd:.3f}'],
                         ["Epochs", f'{epochs}'],
                         ["Learn Rate", f'{learning_rate}'],
                         ["Decay", f'{args.decay}'],
                         ["Batch Size", f'{args.batch}']]
                
                print(tabulate(table, headers="firstrow", tablefmt="fancy_grid", numalign="right"))
                
                cycles += 1
                
                if args.grow:
                    network = grow_network(network,losses,args.grow_thresh,args.grow_width)
                    network = network.to(device)
                
            learning_rate = learning_rate * args.chkdecay 
            filename = f'{args.direct}Metrics/Vae_{model_name}.txt'
            with open(filename, 'w', encoding='utf-8') as file:
                file.write(tabulate(table, headers="firstrow", tablefmt="fancy_grid", numalign="right"))
                
        if args.diffuse:
            cycles = args.cycles
            seg_epochs = epochs//args.noise # segment epoch length for
            for n in range(args.noise+1):             # gradually increasing noise
                print(f'training {n+1} / {args.noise+1}')
                network,losses = trn.train_diffusion(data,network,seg_epochs,learning_rate,wtdecay,
                                                     batch_size,loss_function,print_interval,
                                                     device,n+1,num_points,cycles)
                learning_rate = learning_rate * args.chkdecay
                
            # save network
            torch.save(network.state_dict(),f'{args.direct}Models/diff_{model_name}.pth')
            print(f'Model saved as: {args.direct}Models/diff_{model_name}.pth')
            
            chf_loss,emd = trn.model_eval_chamfer(data, network, num_points, device, args.eval_bs)
            
            # Create and save metric table 
            table = [["Metric", "Value"],
                     ["Chamfer Loss", f'{chf_loss:.3f}'],
                     ["EMD Loss", f'{emd:.3f}'],
                     ["Epochs", f'{epochs}'],
                     ["Learn Rate", f'{learning_rate}'],
                     ["Decay", f'{args.decay}'],
                     ["Batch Size", f'{args.batch}']]
            
            learning_rate = learning_rate * args.chkdecay
            
            print(tabulate(table, headers="firstrow", tablefmt="fancy_grid", numalign="right"))
        
            filename = f'{args.direct}Metrics/diff_{model_name}.txt'
            
            with open(filename, 'w', encoding='utf-8') as file:
                file.write(tabulate(table, headers="firstrow", tablefmt="fancy_grid", numalign="right"))
            
    # GAN training
    if args.gan:
        
        loss_function = trn.GAN_Loss()
        Dnet = networks.jarvis(X[samp].shape[0]).to(device)
        
        if args.loaddis != '':
            if args.loaddis[-4:] != '.pth':
                args.loaddis += '.pth'
            Dnet.load_state_dict(torch.load(f'{args.direct}Models/{args.loaddis}'))
            
        gnet,gloss,dnet,dloss = trn.train_GD(data,network,Dnet,args.hidden1,epochs,
                                             learning_rate,wtdecay,batch_size,
                                             loss_function,print_interval,device)
        
        torch.save(gnet.state_dict(),f'{args.direct}Models/gangen_{model_name}.pth')
        print(f'Generator saved as: {args.direct}Models/gangen_{model_name}.pth')
        
        torch.save(dnet.state_dict(),f'{args.direct}Models/Discrim_{model_name}.pth')
        print(f'Discriminator saved as: {args.direct}Models/Discrim_{model_name}.pth')
        
        chf_loss,emd = trn.model_eval_chamfer(data, gnet, num_points, device, args.eval_bs)
        
        # Create and save metric table 
        table = [["Metric", "Value"],
                 ["Chamfer Loss", f'{chf_loss:.3f}'],
                 ["EMD Loss", f'{emd:.3f}'],
                 ["Epochs", f'{epochs}'],
                 ["Learn Rate", f'{args.lr}'],
                 ["Decay", f'{args.decay}'],
                 ["Batch Size", f'{args.batch}']]
        
        print(tabulate(table, headers="firstrow", tablefmt="fancy_grid", numalign="right"))
        
        filename = f'{args.direct}Metrics/gangen_{model_name}.txt'
        with open(filename, 'w', encoding='utf-8') as file:
            file.write(tabulate(table, headers="firstrow", tablefmt="fancy_grid", numalign="right"))
            
